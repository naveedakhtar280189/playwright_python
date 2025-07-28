import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class ExcelCSVUtils:
    """Reusable utility for reading/writing Excel and CSV files"""

    # ----------- Excel Section -----------
    def read_excel(self, file_path, sheet_name=None, cell=None, row=None, range_str=None):
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name or wb.sheetnames[0]]

            if cell:
                return sheet[cell].value
            elif row:
                return [cell.value for cell in sheet[row]]
            elif range_str:
                return [[cell.value for cell in row] for row in sheet[range_str]]
            else:
                return [[cell.value for cell in row] for row in sheet.iter_rows()]
        except Exception as e:
            print(f"[ERROR] Failed to read Excel file: {e}")
            return None

    def write_excel(self, file_path, sheet_name, cell=None, data=None, row_data=None, col_offset=0):
        try:
            wb = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        if cell and data is not None:
            sheet[cell].value = data
        elif row_data:
            row_num = sheet.max_row + 1
            for idx, value in enumerate(row_data):
                sheet.cell(row=row_num, column=idx + 1 + col_offset, value=value)
        else:
            raise ValueError("Either cell & data or row_data must be provided.")

        try:
            wb.save(file_path)
        except Exception as e:
            print(f"[ERROR] Failed to save Excel file: {e}")

    # ----------- CSV Section -----------
    def read_csv(self, file_path, header=True):
        try:
            return pd.read_csv(file_path) if header else pd.read_csv(file_path, header=None)
        except Exception as e:
            print(f"[ERROR] Failed to read CSV file: {e}")
            return None

    def write_csv(self, file_path, data, header=True):
        try:
            df = pd.DataFrame(data)
            df.to_csv(file_path, index=False, header=header)
        except Exception as e:
            print(f"[ERROR] Failed to write CSV file: {e}")
